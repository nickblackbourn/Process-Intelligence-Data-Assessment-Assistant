"""
Data Ingestion Engine for Process Mining Assessment Tool

This module handles data loading from various file formats with enhanced support for
multi-tab Excel files and embedded schema detection. It provides flexible data
ingestion capabilities optimized for process mining analysis.

Enhanced Features (v2.0):
- Multi-tab Excel processing with automatic tab classification
- Embedded schema detection within Excel sheets
- Flexible CSV encoding and separator detection
- JSON structure analysis and normalization
- Data quality profiling and structure analysis
- Source reference tracking for multi-tab files

Supported Formats:
- CSV files (with encoding/separator auto-detection)
- Excel files (.xlsx, .xls) - single and multi-tab
- JSON files (with structure normalization)
- Parquet files for big data scenarios

Multi-Tab Excel Features:
- Automatic detection of data vs schema definition tabs
- Schema extraction from metadata sheets
- Cross-tab relationship analysis
- Contextual source referencing (file.xlsx#TabName)

Author: Process Intelligence Team
Version: 2.0.0 - Enhanced Multi-Tab Excel Processing
"""

import logging
import os
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
import numpy as np


logger = logging.getLogger(__name__)


class DataIngestionEngine:
    """
    Handles data loading from various file formats with enhanced multi-tab Excel support.
    
    This class provides comprehensive data ingestion capabilities optimized for process
    mining analysis, including advanced Excel processing that can handle complex
    business documents with multiple tabs containing both data and embedded schemas.
    
    Key Capabilities:
    - Multi-format file loading (CSV, Excel, JSON, Parquet)
    - Multi-tab Excel processing with schema detection
    - Automatic data structure analysis and profiling
    - Flexible encoding and format detection
    - Quality assessment and missing data analysis
    
    Multi-Tab Excel Processing:
    - Detects data tabs vs schema definition tabs
    - Extracts embedded schemas from metadata sheets
    - Maintains source references for cross-tab analysis
    - Supports complex enterprise Excel exports
    """
    
    def __init__(self):
        """Initialize the DataIngestionEngine."""
        self.supported_formats = ['.csv', '.xlsx', '.xls', '.json', '.parquet']
        self.excel_multi_tab_enabled = True
        logger.info("DataIngestionEngine initialized")
    
    def load_file(self, file_path: str) -> pd.DataFrame:
        """Load data from various file formats.
        
        Args:
            file_path: Path to the data file
            
        Returns:
            Loaded data as pandas DataFrame
            
        Raises:
            ValueError: If file format is not supported
            FileNotFoundError: If file doesn't exist
        """
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        extension = path.suffix.lower()
        
        if extension not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {extension}")
        
        logger.info(f"Loading data from {file_path}")
        
        try:
            if extension == '.csv':
                # Try different encodings and separators
                data = self._load_csv_flexible(file_path)
            elif extension in ['.xlsx', '.xls']:
                if self.excel_multi_tab_enabled:
                    # Check if file has multiple tabs
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    if len(wb.sheetnames) > 1:
                        logger.info(f"Detected multi-tab Excel file with {len(wb.sheetnames)} sheets")
                        data = self._load_excel_multi_tab(file_path)
                    else:
                        data = pd.read_excel(file_path)
                    wb.close()
                else:
                    data = pd.read_excel(file_path)
            elif extension == '.json':
                data = self._load_json_flexible(file_path)
            elif extension == '.parquet':
                data = pd.read_parquet(file_path)
            else:
                raise ValueError(f"Handler not implemented for {extension}")
            
            # Log success based on data type
            if isinstance(data, dict) and 'multi_tab_file' in data.get('metadata', {}):
                logger.info(f"Successfully loaded multi-tab Excel file with {data['metadata']['total_tabs']} tabs")
            elif hasattr(data, 'shape'):
                logger.info(f"Successfully loaded {len(data)} rows and {len(data.columns)} columns")
            else:
                logger.info(f"Successfully loaded data from {file_path}")
            
            return data
            
        except Exception as e:
            logger.error(f"Error loading data: {e}")
            raise
    
    def _load_csv_flexible(self, file_path: str) -> pd.DataFrame:
        """Load CSV with flexible encoding and separator detection."""
        encodings = ['utf-8', 'latin-1', 'cp1252']
        separators = [',', ';', '\t', '|']
        
        for encoding in encodings:
            for sep in separators:
                try:
                    data = pd.read_csv(file_path, encoding=encoding, sep=sep)
                    if len(data.columns) > 1:  # Must have multiple columns
                        logger.info(f"CSV loaded with encoding={encoding}, separator='{sep}'")
                        return data
                except (UnicodeDecodeError, pd.errors.ParserError):
                    continue
        
        # Fallback to default
        return pd.read_csv(file_path)
    
    def _load_json_flexible(self, file_path: str) -> pd.DataFrame:
        """Load JSON with flexible structure handling."""
        with open(file_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        if isinstance(json_data, list):
            return pd.DataFrame(json_data)
        elif isinstance(json_data, dict):
            # Try to find the main data array
            for key, value in json_data.items():
                if isinstance(value, list) and len(value) > 0:
                    return pd.DataFrame(value)
            # If no array found, treat as single record
            return pd.DataFrame([json_data])
        else:
            raise ValueError("Unsupported JSON structure")
    
    def analyze_structure(self, data: pd.DataFrame) -> Dict[str, Any]:
        """Analyze the structure and characteristics of the dataset.
        
        Args:
            data: Input DataFrame
            
        Returns:
            Dictionary with structural analysis
        """
        analysis = {
            'shape': {
                'rows': len(data),
                'columns': len(data.columns)
            },
            'columns': {},
            'data_types': {},
            'missing_data': {},
            'potential_identifiers': [],
            'potential_timestamps': [],
            'potential_activities': [],
            'data_quality_indicators': {}
        }
        
        # Analyze each column
        for col in data.columns:
            col_analysis = self._analyze_column(data, col)
            analysis['columns'][col] = col_analysis
            
            # Categorize columns based on characteristics
            if col_analysis['could_be_identifier']:
                analysis['potential_identifiers'].append(col)
            
            if col_analysis['could_be_timestamp']:
                analysis['potential_timestamps'].append(col)
                
            if col_analysis['could_be_activity']:
                analysis['potential_activities'].append(col)
        
        # Data type distribution
        type_counts = data.dtypes.value_counts()
        analysis['data_types'] = {str(dtype): int(count) for dtype, count in type_counts.items()}
        
        # Missing data analysis
        missing_counts = data.isnull().sum()
        analysis['missing_data'] = {
            'total_missing': int(missing_counts.sum()),
            'missing_percentage': float((missing_counts.sum() / (len(data) * len(data.columns))) * 100),
            'columns_with_missing': [col for col in data.columns if data[col].isnull().any()]
        }
        
        # Data quality indicators
        analysis['data_quality_indicators'] = {
            'completeness': float(100 - analysis['missing_data']['missing_percentage']),
            'duplicate_rows': int(data.duplicated().sum()),
            'duplicate_percentage': float((data.duplicated().sum() / len(data)) * 100),
            'empty_strings': self._count_empty_strings(data),
            'potential_data_issues': self._identify_data_issues(data)
        }
        
        return analysis
    
    def _analyze_column(self, data: pd.DataFrame, col: str) -> Dict[str, Any]:
        """Analyze individual column characteristics."""
        col_data = data[col]
        
        analysis = {
            'dtype': str(col_data.dtype),
            'unique_count': col_data.nunique(),
            'unique_percentage': (col_data.nunique() / len(col_data)) * 100,
            'missing_count': col_data.isnull().sum(),
            'missing_percentage': (col_data.isnull().sum() / len(col_data)) * 100,
            'could_be_identifier': False,
            'could_be_timestamp': False,
            'could_be_activity': False,
            'sample_values': []
        }
        
        # Get sample values (non-null)
        non_null_values = col_data.dropna()
        if len(non_null_values) > 0:
            sample_size = min(5, len(non_null_values))
            analysis['sample_values'] = non_null_values.head(sample_size).tolist()
        
        # Check if could be identifier (but exclude timestamp-like columns)
        timestamp_keywords = ['time', 'date', 'timestamp', 'created', 'updated', 'occurred']
        is_timestamp_like = any(keyword in col.lower() for keyword in timestamp_keywords)
        
        # For case ID detection, also check column name patterns
        case_id_keywords = ['id', 'case', 'order', 'ticket', 'instance', 'key']
        has_id_pattern = any(keyword in col.lower() for keyword in case_id_keywords)
        
        # Relaxed case ID detection - allow lower uniqueness if name pattern matches
        if has_id_pattern and analysis['missing_percentage'] < 10 and not is_timestamp_like:
            analysis['could_be_identifier'] = True
        elif (analysis['unique_percentage'] > 80 and 
              analysis['missing_percentage'] < 10 and 
              not is_timestamp_like):
            analysis['could_be_identifier'] = True
        
        # Check if could be timestamp
        if col_data.dtype == 'object':
            # Try to parse as datetime
            try:
                pd.to_datetime(non_null_values.head(100), errors='raise')
                analysis['could_be_timestamp'] = True
            except:
                pass
            
            # Check for activity-like patterns
            activity_name_keywords = ['status', 'state', 'activity', 'event', 'action', 'step', 'phase', 'stage']
            has_activity_name = any(keyword in col.lower() for keyword in activity_name_keywords)
            
            # Activity detection: reasonable number of unique values + activity-like name or content
            if (analysis['unique_count'] >= 3 and 
                analysis['unique_count'] <= len(col_data) * 0.8 and 
                analysis['missing_percentage'] < 20):
                
                if has_activity_name:
                    analysis['could_be_activity'] = True
                else:
                    # Check if values look like activities
                    sample_str = str(non_null_values.iloc[0]).lower() if len(non_null_values) > 0 else ""
                    activity_content_keywords = ['create', 'process', 'send', 'receive', 'approve', 'reject', 
                                               'start', 'end', 'complete', 'cancel', 'update', 'delete',
                                               'submit', 'review', 'confirm', 'ship', 'deliver']
                    if any(keyword in sample_str for keyword in activity_content_keywords):
                        analysis['could_be_activity'] = True
        
        # Additional numeric analysis
        if col_data.dtype in ['int64', 'float64']:
            numeric_data = col_data.dropna()
            if len(numeric_data) > 0:
                std_val = numeric_data.std() if len(numeric_data) > 1 else 0
                analysis.update({
                    'mean': float(numeric_data.mean()),
                    'median': float(numeric_data.median()),
                    'std': float(std_val) if not pd.isna(std_val) else 0.0,
                    'min': float(numeric_data.min()),
                    'max': float(numeric_data.max())
                })
        
        return analysis
    
    def _count_empty_strings(self, data: pd.DataFrame) -> int:
        """Count empty strings across all object columns."""
        empty_count = 0
        for col in data.select_dtypes(include=['object']).columns:
            empty_count += (data[col].astype(str).str.strip() == '').sum()
        return int(empty_count)
    
    def _identify_data_issues(self, data: pd.DataFrame) -> List[str]:
        """Identify potential data quality issues."""
        issues = []
        
        # Check for columns with very high missing data
        for col in data.columns:
            missing_pct = (data[col].isnull().sum() / len(data)) * 100
            if missing_pct > 50:
                issues.append(f"Column '{col}' has {missing_pct:.1f}% missing data")
        
        # Check for columns with very low variability
        for col in data.select_dtypes(include=['object']).columns:
            if data[col].nunique() == 1:
                issues.append(f"Column '{col}' has only one unique value")
        
        # Check for suspicious patterns
        if len(data) < 10:
            issues.append("Dataset is very small (less than 10 rows)")
        
        if len(data.columns) < 3:
            issues.append("Dataset has very few columns for process mining")
        
        return issues
    
    def clean_data(self, data: pd.DataFrame) -> pd.DataFrame:
        """Clean and preprocess the data for analysis.
        
        Args:
            data: Raw data DataFrame
            
        Returns:
            Cleaned data DataFrame
        """
        logger.info("Starting data cleaning")
        
        # Make a copy to avoid modifying original data
        cleaned_data = data.copy()
        
        # Remove completely empty rows and columns
        cleaned_data = cleaned_data.dropna(how='all', axis=0)
        cleaned_data = cleaned_data.dropna(how='all', axis=1)
        
        # Strip whitespace from string columns
        string_cols = cleaned_data.select_dtypes(include=['object']).columns
        for col in string_cols:
            cleaned_data[col] = cleaned_data[col].astype(str).str.strip()
            # Replace empty strings with NaN
            cleaned_data[col] = cleaned_data[col].replace('', pd.NA)
        
        # Convert potential numeric columns
        for col in cleaned_data.columns:
            if cleaned_data[col].dtype == 'object':
                # Try to convert to numeric
                numeric_series = pd.to_numeric(cleaned_data[col], errors='ignore')
                if not numeric_series.equals(cleaned_data[col]):
                    cleaned_data[col] = numeric_series
        
        logger.info(f"Data cleaning completed. Shape: {cleaned_data.shape}")
        return cleaned_data

    def _load_excel_multi_tab(self, file_path: str) -> Dict[str, Any]:
        """Load Excel file with multi-tab processing and schema detection.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary containing all tabs with metadata and potential schemas
        """
        import openpyxl
        
        result = {
            'file_path': file_path,
            'tabs': {},
            'metadata': {
                'total_tabs': 0,
                'data_tabs': [],
                'schema_tabs': [],
                'empty_tabs': [],
                'multi_tab_file': True
            }
        }
        
        try:
            # Load workbook to get all sheet names
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            result['metadata']['total_tabs'] = len(sheet_names)
            
            logger.info(f"Processing Excel file with {len(sheet_names)} tabs: {sheet_names}")
            
            for sheet_name in sheet_names:
                try:
                    # Read each sheet
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    if df.empty or len(df.columns) == 0:
                        result['metadata']['empty_tabs'].append(sheet_name)
                        continue
                    
                    # Analyze sheet content for schema vs data
                    sheet_analysis = self._analyze_excel_sheet_content(df, sheet_name)
                    
                    if sheet_analysis['is_schema_definition']:
                        result['metadata']['schema_tabs'].append(sheet_name)
                    else:
                        result['metadata']['data_tabs'].append(sheet_name)
                    
                    result['tabs'][sheet_name] = {
                        'data': df,
                        'analysis': sheet_analysis,
                        'metadata': self.analyze_structure(df),
                        'source_reference': f"{file_path}#{sheet_name}"
                    }
                    
                    logger.info(f"Processed sheet '{sheet_name}': {len(df)} rows, {len(df.columns)} columns")
                    
                except Exception as e:
                    logger.warning(f"Failed to process sheet '{sheet_name}': {e}")
                    continue
            
            workbook.close()
            
        except Exception as e:
            logger.error(f"Failed to load Excel file {file_path}: {e}")
            raise
        
        return result

    def _analyze_excel_sheet_content(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Analyze if Excel sheet contains schema definition or data.
        
        Args:
            df: DataFrame containing sheet data
            sheet_name: Name of the Excel sheet
            
        Returns:
            Analysis results including schema detection
        """
        analysis = {
            'is_schema_definition': False,
            'is_data_table': False,
            'schema_type': None,
            'schema_elements': [],
            'confidence': 0.0,
            'sheet_name': sheet_name
        }
        
        # Check for schema definition patterns in column names
        schema_indicators = {
            'data_dictionary': ['table_name', 'column_name', 'data_type', 'description', 'nullable', 'field', 'type'],
            'entity_relationship': ['entity', 'attribute', 'relationship', 'cardinality', 'table', 'key'],
            'process_mapping': ['process', 'activity', 'step', 'input', 'output', 'role', 'actor'],
            'lookup_table': ['code', 'value', 'description', 'category', 'status'],
            'configuration': ['parameter', 'setting', 'config', 'property', 'name', 'value']
        }
        
        # Normalize column names for comparison
        column_names = [col.lower().replace(' ', '_').replace('-', '_') for col in df.columns]
        
        # Check sheet name patterns
        schema_sheet_names = ['schema', 'data_dictionary', 'metadata', 'config', 'mapping', 'lookup', 'reference']
        if any(pattern in sheet_name.lower() for pattern in schema_sheet_names):
            analysis['confidence'] += 0.3
        
        # Check column patterns
        best_match_score = 0
        best_match_type = None
        
        for schema_type, indicators in schema_indicators.items():
            matches = sum(1 for indicator in indicators if any(indicator in col for col in column_names))
            confidence = matches / len(indicators) if indicators else 0
            
            if confidence > best_match_score:
                best_match_score = confidence
                best_match_type = schema_type
        
        # Combine confidence scores
        total_confidence = (analysis['confidence'] + best_match_score) / 2
        
        if total_confidence > 0.4 or best_match_score > 0.6:  # Schema definition threshold
            analysis['is_schema_definition'] = True
            analysis['schema_type'] = best_match_type
            analysis['confidence'] = total_confidence
            analysis['schema_elements'] = self._extract_excel_schema_elements(df, best_match_type)
        else:
            analysis['is_data_table'] = True
        
        return analysis

    def _extract_excel_schema_elements(self, df: pd.DataFrame, schema_type: str) -> List[Dict[str, Any]]:
        """Extract schema elements from Excel schema definition sheets.
        
        Args:
            df: DataFrame containing schema information
            schema_type: Type of schema detected
            
        Returns:
            List of schema elements
        """
        elements = []
        
        # Map common column patterns to standardized names
        column_mapping = {
            'data_dictionary': {
                'name_cols': ['column_name', 'field_name', 'field', 'name', 'attribute'],
                'type_cols': ['data_type', 'type', 'dtype', 'column_type'],
                'desc_cols': ['description', 'desc', 'comment', 'notes'],
                'table_cols': ['table_name', 'table', 'entity', 'source']
            },
            'process_mapping': {
                'name_cols': ['activity', 'step', 'task', 'process_step', 'name'],
                'type_cols': ['type', 'category', 'phase'],
                'desc_cols': ['description', 'desc', 'details'],
                'input_cols': ['input', 'inputs', 'predecessor'],
                'output_cols': ['output', 'outputs', 'successor']
            }
        }
        
        if schema_type in column_mapping:
            mapping = column_mapping[schema_type]
            df_cols = [col.lower().replace(' ', '_') for col in df.columns]
            
            # Find matching columns
            name_col = self._find_matching_column(df_cols, mapping['name_cols'])
            type_col = self._find_matching_column(df_cols, mapping.get('type_cols', []))
            desc_col = self._find_matching_column(df_cols, mapping.get('desc_cols', []))
            
            if name_col:
                original_name_col = df.columns[df_cols.index(name_col)]
                
                for _, row in df.iterrows():
                    if pd.notna(row.get(original_name_col)):
                        element = {
                            'name': str(row.get(original_name_col, '')),
                            'source': 'excel_schema_definition',
                            'schema_type': schema_type
                        }
                        
                        if type_col:
                            original_type_col = df.columns[df_cols.index(type_col)]
                            element['type'] = str(row.get(original_type_col, ''))
                        
                        if desc_col:
                            original_desc_col = df.columns[df_cols.index(desc_col)]
                            element['description'] = str(row.get(original_desc_col, ''))
                        
                        # Add any additional fields
                        for key, value in row.items():
                            if key not in [original_name_col, type_col, desc_col] and pd.notna(value):
                                element[key.lower().replace(' ', '_')] = str(value)
                        
                        elements.append(element)
        
        return elements

    def _find_matching_column(self, df_columns: List[str], target_patterns: List[str]) -> Optional[str]:
        """Find the first column that matches any of the target patterns.
        
        Args:
            df_columns: List of normalized column names
            target_patterns: List of patterns to match
            
        Returns:
            First matching column name or None
        """
        for pattern in target_patterns:
            for col in df_columns:
                if pattern in col:
                    return col
        return None
    
    def export_to_yaml(self, data: Dict[str, Any], file_path: str) -> None:
        """Export analysis results to YAML file.
        
        Args:
            data: Data to export
            file_path: Output file path
        """
        import yaml
        
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                yaml.dump(data, f, default_flow_style=False, sort_keys=False)
            
            logger.info(f"Data exported to YAML: {file_path}")
            
        except Exception as e:
            logger.error(f"Error exporting to YAML: {e}")
            raise
